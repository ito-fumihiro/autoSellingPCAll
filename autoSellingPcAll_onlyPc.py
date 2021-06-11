from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import urllib.request
from selenium.webdriver.support.ui import Select
import openpyxl
import traceback
import logging
import sys
import random
import datetime
from bs4 import BeautifulSoup
import re
import pyperclip
import glob
import pywinauto
import shutil

try:

	# 実行パスを変更
	dirPath = sys.argv[1]
	os.chdir(dirPath)

	# HTMLファイルのパス
	htmlFilePath = "html\\" +sys.argv[2]

	# TEXTファイルのパス
	textFilePath = "text\\" +sys.argv[3]

	# 出品ボタンを押すか
	isExhibitString = sys.argv[4]

	# 出品待機時間
	exhibitWaitTime = int(float(sys.argv[5]) * 3600)

	LISTED_DIR = 'input_data\\出品済み'

	# 出品後フォルダの作成
	if not os.path.exists(LISTED_DIR):
		# ディレクトリが存在しない場合、ディレクトリを作成する
		os.makedirs(LISTED_DIR)


	# ログ設定
	logger = logging.getLogger('logging_auto_selling')
	logFilePath = os.getcwd() + "\\logging_auto_selling.log"
	fh = logging.FileHandler(logFilePath)  # ファイル名を設定
	logger.addHandler(fh)
	logger.setLevel(10)

	logger.info("arg[1]:" + sys.argv[1])
	logger.info("arg[2]:" + sys.argv[2])
	logger.info("arg[3]:" + sys.argv[3])
	logger.info("arg[4]:" + sys.argv[4])
	logger.info("arg[5]:" + sys.argv[5])

	tomorrow = datetime.date.today() + datetime.timedelta(days=1)

	#リターンコード
	returnCode = 0
	logger.info("Auto Selling Pc All Start")

	#Chromeの設定
	userdata_dir = dirPath + '/UserData'  # カレントディレクトリの直下に作る場合
	os.makedirs(userdata_dir, exist_ok=True)
	options = webdriver.ChromeOptions()
	options.add_argument('--user-data-dir=' + userdata_dir)
	driver = webdriver.Chrome(options=options)

	#エクセルファイルを開く
	wb = openpyxl.load_workbook('GoodsList.xlsm', data_only=True)

	merucariSheet = wb.get_sheet_by_name('メルカリ')
	rakumaSheet = wb.get_sheet_by_name('ラクマ')
	yafuokuSheet = wb.get_sheet_by_name('ヤフオク')
	jmtySheet = wb.get_sheet_by_name('ジモティー')
	goodsSheet = wb.get_sheet_by_name('コピー')

	yafuokuCategorySheet = wb.get_sheet_by_name('カテゴリ対応表_ヤフオク')


	rc = 2
	count = 0
	w_list = ['(月)', '(火)', '(水)', '(木)', '(金)', '(土)', '(日)']
	while rc <= goodsSheet.max_row:

		count = count +1

		# コピーシートの値を取得
		#商品名(ヤフオク、ジモティ用)
		goodsName = goodsSheet.cell(row=rc, column=3).value

		#商品名(メルカリ、ラクマ用)
		goodsName2 = goodsSheet.cell(row=rc, column=26).value

		if not goodsName and not goodsName2:
			rc = rc +1
			continue

		# ファイルをオープンする
		html_data = open(htmlFilePath, "r", encoding="utf-8")
		text_data = open(textFilePath, "r", encoding="utf-8")
		htmlContents = html_data.read()
		textContents = text_data.read()


		#プレースホルダーを書き換える
		htmlContents = htmlContents.replace("$title$", str(goodsName)) 
		textContents = textContents.replace("$title$", str(goodsName)) 

		# メーカー			
		pres_maker = goodsSheet.cell(row=rc, column=6).value
		htmlContents = htmlContents.replace("$maker$", str(pres_maker))
		textContents = textContents.replace("$maker$", str(pres_maker))

		# 型番
		pres_typeNumber = goodsSheet.cell(row=rc, column=7).value
		htmlContents = htmlContents.replace("$typeNumber$", str(pres_typeNumber))
		textContents = textContents.replace("$typeNumber$", str(pres_typeNumber))

		# CPU
		pres_cpu = goodsSheet.cell(row=rc, column=8).value
		htmlContents = htmlContents.replace("$cpu$", str(pres_cpu))
		textContents = textContents.replace("$cpu$", str(pres_cpu))

		# メモリー
		pres_memory = goodsSheet.cell(row=rc, column=9).value
		htmlContents = htmlContents.replace("$memory$", str(pres_memory))
		textContents = textContents.replace("$memory$", str(pres_memory))

		# 容量
		pres_volume = goodsSheet.cell(row=rc, column=10).value
		htmlContents = htmlContents.replace("$volume$", str(pres_volume))
		textContents = textContents.replace("$volume$", str(pres_volume))

		# ドライブ
		pres_drive = goodsSheet.cell(row=rc, column=13).value
		htmlContents = htmlContents.replace("$drive$", str(pres_drive))
		textContents = textContents.replace("$drive$", str(pres_drive))

		# アクセサリー
		pres_accessory = goodsSheet.cell(row=rc, column=18).value
		htmlContents = htmlContents.replace("$accessory$", str(pres_accessory))
		textContents = textContents.replace("$accessory$", str(pres_accessory))

		# 在庫番号
		stockNum = goodsSheet.cell(row=rc, column=1).value
		htmlContents = htmlContents.replace("$stockNum$", str(stockNum))
		textContents = textContents.replace("$stockNum$", str(stockNum))

		# ランク
		pres_rank = goodsSheet.cell(row=rc, column=4).value
		htmlContents = htmlContents.replace("$rank$", str(pres_rank))
		textContents = textContents.replace("$rank$", str(pres_rank))

		# カメラ
		pres_camera = goodsSheet.cell(row=rc, column=12).value
		htmlContents = htmlContents.replace("$camera$", str(pres_camera))
		textContents = textContents.replace("$camera$", str(pres_camera))

		# バッテリー
		pres_battery = goodsSheet.cell(row=rc, column=15).value
		htmlContents = htmlContents.replace("$battery$", str(pres_battery))
		textContents = textContents.replace("$battery$", str(pres_battery))

		# 液晶ディスプレー
		pres_lcd = goodsSheet.cell(row=rc, column=17).value
		htmlContents = htmlContents.replace("$lcd$", str(pres_lcd))
		textContents = textContents.replace("$lcd$", str(pres_lcd))

		# その他
		pres_other = goodsSheet.cell(row=rc, column=5).value
		htmlContents = htmlContents.replace("$other$", str(pres_other))
		textContents = textContents.replace("$other$", str(pres_other))

		# ファイルをクローズする
		html_data.close()
		text_data.close()
		description_html = htmlContents
		description_text = textContents

		#送料を入力するかどうか(独自項目)
		isSetFee = goodsSheet.cell(row=rc, column=23).value

		# 2020/12/07
		# 開始価格、即決価格入力
		auctionStartPrice = int(goodsSheet.cell(row=rc, column=24).value)
		auctionDecidePrice = int(goodsSheet.cell(row=rc, column=25).value)

		# メルカリ----------------------
		try:
			isSellingMerucari = True
			logger.info("merucari" + str(count))
			driver.execute_script("window.open()") #新しいタブを開く
			# ウィンドウハンドルを取得する
			handle_array = driver.window_handles

			# 一番最後のdriverに切り替える
			driver.switch_to.window(handle_array[len(handle_array)-1])

			# メルカリにアクセス
			driver.get('https://www.mercari.com/jp/sell/')
			#指定されたフレームが利用出来るまで待機する
			WebDriverWait(driver, 500).until(
				EC.presence_of_element_located((By.NAME, "description"))
			)

			# 読み込み遅いかもしれないから3秒待つ。
			sleep(3)

			#画像
			folderPath = "input_data/image/" + str(goodsSheet.cell(row=rc, column=22).value)
			files = glob.glob(folderPath + "/*")
			imgCount = 0
			for file_name in files:								
				if os.path.exists(file_name):
					#driver.find_element_by_xpath("//input[@type='file']").find_element(By.XPATH,'../..').click()
					JavaScriptFocusToElement = "arguments[0].focus()"
					element = driver.find_element_by_xpath("//input[@type='file']").find_element(By.XPATH,'../..')
					driver.execute_script(JavaScriptFocusToElement, element)
					element.send_keys(Keys.ENTER)
					sleep(2)

					# 開くダイアログを探して接続する
					findWindow = lambda: pywinauto.findwindows.find_windows(title=u'開く')[0]
					dialog = pywinauto.timings.wait_until_passes(5, 1, findWindow)

					# pywinauto に探し出したダイアログを接続
					pwa_app = pywinauto.Application()
					pwa_app.connect(handle=dialog)
					window = pwa_app[u"開く"]

					addres = window.children()[39]
					addres.click()

					dialog_dir = window.children()[43]
					dialog_dir.type_keys( dirPath +"\\"+ folderPath+'{ENTER}',with_spaces=True)

					filePath = file_name.rsplit('\\', 1)[1]
					# テキストボックス(ファイル名)にPATHを入力
					tb = window[u"ファイル名(&N):"]
					if tb.is_enabled():
						tb.click()
						edit = window.Edit4
						edit.set_focus()
						# ファイルを選択し、Alt + Oを押下
						edit.type_keys(filePath + '%O',with_spaces=True)

					sleep(3)
					imgCount = imgCount +1   
					if imgCount >= 10:
						break 

			# エクセルファイルから値を取得
			category1 = "家電・スマホ・カメラ"
			category2 = "PC/タブレット"
			category3 = "デスクトップ型PC"
			if goodsSheet.cell(row=rc, column=21).value == "ノートブック、ノートパソコン":
				category3 = "ノートPC"
			category4 = "-"
			category5 = "-"

			size = merucariSheet.cell(row=7, column=20).value 
			brand = merucariSheet.cell(row=7, column=21).value 
			status = merucariSheet.cell(row=7, column=22).value 
			deliveryBurden = merucariSheet.cell(row=7, column=23).value 
			deliveryMethod = merucariSheet.cell(row=7, column=24).value
			destinationRegion = merucariSheet.cell(row=7, column=25).value 
			dayToSend = merucariSheet.cell(row=7, column=26).value 

			# 商品名
			driver.find_element_by_xpath( "//input[@name='name']" ).send_keys(goodsName2)
			sleep(2)

			# 商品の説明
			pyperclip.copy(description_text)
			driver.find_element_by_xpath( "//textarea[@name='description']" ).send_keys(Keys.CONTROL+ "v")
			sleep(2)

			#カテゴリー

			Select(driver.find_elements_by_xpath( "//select[@name='categoryId']" )[0]).select_by_visible_text(category1)
			sleep(2)
			Select(driver.find_elements_by_xpath( "//select[@name='categoryId']" )[1]).select_by_visible_text(category2)
			sleep(2)
			Select(driver.find_elements_by_xpath( "//select[@name='categoryId']" )[2]).select_by_visible_text(category3)
			sleep(2)


			#サイズ

			#ブランド

			#商品の状態
			if  status == "-" or  not status:
				print("商品の状態をスキップ")
			else:
				Select(driver.find_element_by_xpath("//select[@name='itemCondition']")).select_by_visible_text(status)
				sleep(2)

			#配送料の負担
			if deliveryBurden == "-" or not deliveryBurden:	
				print("配送料の負担をスキップ")
			else:						
				Select(driver.find_element(By.XPATH, '//select[@name="shippingPayer"]')).select_by_visible_text(deliveryBurden)
				sleep(2)

			#配送の方法
			if deliveryMethod == "-" or not deliveryMethod:
				print("配送の方法をスキップ")
			else:
				Select(driver.find_element(By.XPATH, '//select[@name="shippingMethod"]')).select_by_visible_text(deliveryMethod)
				sleep(2)

			#配送の地域
			if destinationRegion == "-" or not destinationRegion:	
				print("配送の地域をスキップ")
			else:
				Select(driver.find_element(By.XPATH, '//select[@name="shippingFromArea"]')).select_by_visible_text(destinationRegion)
				sleep(2)

			#発送までの日数
			if dayToSend == "-" or  not dayToSend:
				print("発送までの日数をスキップ")
			else:
				Select(driver.find_element(By.XPATH, '//select[@name="shippingDuration"]')).select_by_visible_text(dayToSend)
				sleep(2)

			#価格
			if auctionStartPrice == "-" or not auctionStartPrice:
				print("価格をスキップ")
			else:
				driver.find_element_by_xpath( "//input[@name='price']" ).send_keys(auctionStartPrice)
			sleep(2)

			#出品するかしないか
			if isExhibitString == "T":
				xpath = '//button[@type="submit"]'
				element = driver.find_element(By.XPATH, xpath)
				element.send_keys(Keys.ENTER)

		except Exception as merucari_err:
			logger.exception('Raise Exception merucari: %s', merucari_err)

		# ----------------------メルカリ

		# ラクマ----------------------
		try:
			logger.info("rakuma" + str(count))
			isSellingRakuma = True
			driver.execute_script("window.open()") #新しいタブを開く
			# ウィンドウハンドルを取得する
			handle_array = driver.window_handles

			# 一番最後のdriverに切り替える
			driver.switch_to.window(handle_array[len(handle_array)-1])
			driver.get('https://fril.jp/item/new')

			#指定されたフレームが利用出来るまで待機する
			WebDriverWait(driver, 500).until(
				EC.presence_of_element_located((By.ID, "detail"))
			)

			import glob
			folderPath = "input_data/image/" + str(goodsSheet.cell(row=rc, column=22).value) + "/*" 
			files = glob.glob(folderPath)
			imageCount = 0
			for file_name in files:
				imageName = file_name
				if os.path.exists(imageName):
					driver.find_elements_by_xpath( "//input[@type='file']")[imageCount].send_keys(os.path.abspath(imageName))
					imageCount = imageCount +1
					sleep(3)
					if imageCount >= 4:
						break	
				else:
					continue

			# エクセルファイルから値を取得
			size = rakumaSheet.cell(row=7, column=12).value 
			brand = rakumaSheet.cell(row=7, column=13).value 
			status = rakumaSheet.cell(row=7, column=14).value 
			deliveryBurden = rakumaSheet.cell(row=7, column=15).value 
			deliveryMethod = rakumaSheet.cell(row=7, column=16).value
			dayToSend = rakumaSheet.cell(row=7, column=17).value 
			destinationRegion = rakumaSheet.cell(row=7, column=18).value 
			PurchaseApplication = rakumaSheet.cell(row=7, column=19).value

			#カテゴリ
			category1 = "スマホ/家電/カメラ"
			category2 = "PC/タブレット"
			# カテゴリ３はデスクトップかノートパソコンかの分岐しかない
			category3 = "デスクトップ型PC"
			if goodsSheet.cell(row=rc, column=21).value == "ノートブック、ノートパソコン":
				category3 = "ノートPC"

			categorylist = []
			categorylist.append(category1)
			categorylist.append(category2)
			categorylist.append(category3)

			#商品名
			if goodsName2 == None:
				print("商品名をスキップ")
			else:
				driver.find_element_by_xpath( "//input[@id='name']" ).send_keys(goodsName2)
				sleep(2)

			#商品説明
			if description_text == None:
				print("商品説明をスキップ")
			else:
				pyperclip.copy(description_text)
				driver.find_element_by_xpath( "//textarea[@id='detail']" ).send_keys(Keys.CONTROL+ "v")
				sleep(2)    

			#カテゴリ
			driver.find_element_by_id('category_name').click()
			sleep(2)


			#カテゴリー
			cc  = 0
			nowElement = driver.find_element_by_id('select-category')
			while cc < 3:

				category = categorylist[cc]
				xpath = '//*[text()="%s"]' % category
				cc= cc+1
				elements = nowElement.find_elements(By.XPATH, xpath)
				for item in elements:
					parent = item.find_element(By.XPATH,'..')
					if cc == 1:
						if parent.get_attribute("class") == 'list-group-item parent':										
							parent.click()
							sleep(2)
							break
					elif cc == 2:
						if item.get_attribute("class") == 'list-group-item small branch' :
							item.click()
							sleep(2)
							break
					elif cc == 3:
						if item.get_attribute("class") == 'list-group-item small leaf' :
							item.click()
							sleep(2)
							break

			#サイズ
			if  size == "-" or  not size:
				print("サイズをスキップ")
			else:
				driver.find_element_by_id('size_name').click()
				sleep(2)
				xpath = '//a[text()="%s"]' % size
				driver.find_element(By.XPATH, xpath).click()
				sleep(2)

			#ブランド
			if brand == "-" or  not brand:
				print("ブランドをスキップ")
			else:
				driver.find_element_by_id('brand_name').click()
				sleep(2)
				driver.find_element_by_id('brand-search-text').send_keys(" ")
				sleep(2)
				xpath = '//span[text()="%s"]' % brand
				driver.find_element(By.XPATH, xpath).find_element(By.XPATH,'..').click()
				sleep(2)

			#商品の状態
			Select(driver.find_element_by_xpath( "//select[@id='status']" )).select_by_visible_text(status)
			sleep(2)

			#配送の負担
			Select(driver.find_element_by_xpath( "//select[@id='carriage']" )).select_by_visible_text(deliveryBurden)
			sleep(2)

			#配送方法
			driver.find_element_by_id('delivery_method').click()
			sleep(2)
			xpath = '//span[text()="%s"]' % deliveryMethod
			driver.find_element(By.XPATH, xpath).click()
			sleep(2)

			#発送日の目安
			Select(driver.find_element_by_xpath( "//select[@id='delivery_date']" )).select_by_visible_text(dayToSend)
			sleep(2)

			#発送元の地域
			Select(driver.find_element_by_xpath( "//select[@id='delivery_area']" )).select_by_visible_text(destinationRegion)
			sleep(2)

			#購入申請
			Select(driver.find_element_by_xpath( "//select[@id='request_required']" )).select_by_visible_text(PurchaseApplication)
			sleep(2)

			#商品価格
			if auctionStartPrice == None:
				print("商品価格をスキップ")
			else:
				driver.find_element_by_xpath( "//input[@id='sell_price']" ).send_keys(auctionStartPrice)
				sleep(2)

			#確認画面へ
			driver.find_element_by_xpath( "//*[@id='confirm']" ).click()
			sleep(2)

			#出品するかしないか
			if isExhibitString == "T":
				driver.find_element_by_xpath( "//*[@id='submit']" ).click()
				sleep(2)

		except Exception as rakuma_err:
			logger.exception('Raise Exception rakuma: %s', rakuma_err)
		#----------------------ラクマ

		# 処理開始 ----------------------ヤフオク
		try:
			logger.info("Yafuoku" + str(count))
			driver.execute_script("window.open()")  # 新しいタブを開く
			# ウィンドウハンドルを取得する
			handle_array = driver.window_handles

			# 一番最後のdriverに切り替える
			driver.switch_to.window(handle_array[len(handle_array)-1])

			driver.get('https://auctions.yahoo.co.jp/sell/jp/show/submit?category=0')
			# 読み込み遅いかもしれないから3秒待つ。
			sleep(2)

			if not goodsSheet.cell(row=rc, column=22).value or yafuokuSheet.cell(row=rc, column=22).value == "0":
				pass
			else:
				import glob
				folderPath = "input_data/image/" + str(goodsSheet.cell(row=rc, column=22).value) + "/*" 
				files = glob.glob(folderPath)
				cc = 0
				for file_name in files:
					cc = cc +1
					imageName = file_name
					if os.path.exists(imageName):
						driver.find_element_by_xpath("//input[@type='file']").send_keys(os.path.abspath(imageName))
						sleep(3)
					else:
						continue

			#ブランド(固定)
			brand = yafuokuSheet.cell(row=7, column=24).value
			#サイズ(固定)
			size = yafuokuSheet.cell(row=7, column=25).value
			#商品の状態(固定)
			status = yafuokuSheet.cell(row=7, column=26).value

			#返品を受け取る(固定)
			getReturn = yafuokuSheet.cell(row=7, column=28).value
			#発送元の地域(固定)
			destinationRegion = yafuokuSheet.cell(row=7, column=29).value
			#送料負担(固定)
			postageBurdenPerson = yafuokuSheet.cell(row=7, column=30).value
			#落札された後に送料を連絡する(固定)
			shippingCostWay = yafuokuSheet.cell(row=7, column=31).value
			#ヤフネコ！ネコポス(固定)
			yafunekoNekopos = yafuokuSheet.cell(row=7, column=32).value
			#ヤフネコ！宅配便コンパクト(固定)
			yafunekoTakuConpact = yafuokuSheet.cell(row=7, column=33).value
			#ヤフネコ！宅配便(固定)
			yafunekoTaku = yafuokuSheet.cell(row=7, column=34).value
			#ヤフネコ！宅配便　縦、横、高さの合計(固定)
			yafunekoTakuSize = yafuokuSheet.cell(row=7, column=35).value
			#ヤフネコ！宅配便　重さ(固定)
			yafunekoTakuWeight = yafuokuSheet.cell(row=7, column=36).value
			#ゆうパケット(おてがる版)(固定)
			yupaketOtegaru = yafuokuSheet.cell(row=7, column=37).value
			#ゆうパック(おてがる版)(固定)
			yupackOtegaru = yafuokuSheet.cell(row=7, column=38).value
			#ゆうパック(おてがる版)　縦、横、高さの合計(固定)
			yupackSize = yafuokuSheet.cell(row=7, column=39).value
			#ゆうパック(おてがる版)　重さ(固定)
			yupackWeight = yafuokuSheet.cell(row=7, column=40).value
			#その他の配送方法１(固定)
			sendMethodOther1 = yafuokuSheet.cell(row=7, column=41).value
			#その他の配送方法2(固定)
			sendMethodOther2 = yafuokuSheet.cell(row=7, column=42).value
			#支払いから発送までの日数(固定)
			sendToDate = yafuokuSheet.cell(row=7, column=43).value
			#販売形式 オークション(固定)
			sellingAuction = yafuokuSheet.cell(row=7, column=44).value
			#販売形式 定額(固定)
			sellingFixed = yafuokuSheet.cell(row=7, column=45).value

			#オークション開始価格(固定) → 固定解除
			#auctionStartPrice = sheet.cell(row=7, column=46).value
			#オークション即決価格(固定) → 固定解除
			#auctionDecidePrice = sheet.cell(row=7, column=47).value

			#固定価格(固定)
			fixedPrice = yafuokuSheet.cell(row=7, column=48).value	
			#値下げ交渉
			reductionNego = yafuokuSheet.cell(row=7, column=49).value

			# 終了日時は 1日後とする
			date = datetime.date.today() + datetime.timedelta(days=1)
			month = str(date.month) + "月"
			day = str(date.day) + "日"
			weekday = w_list[date.weekday()]
			endDay = str(month) + str(day) + " " + weekday


			#終了時間(ランダム) 19時～23時
			rand = random.randint(19,23)
			endHour = str(rand) + "時台"

			#自動延長
			autoPostponed = yafuokuSheet.cell(row=7, column=52).value
			#自動再出品回数
			autoReSellingCount = yafuokuSheet.cell(row=7, column=53).value				
			#自動再出品値下げ
			autoReSellingReduce = yafuokuSheet.cell(row=7, column=54).value
			#総合評価で制限する
			limitTotalEval = yafuokuSheet.cell(row=7, column=55).value
			#悪い評価の割合で制限する
			limitBadEval = yafuokuSheet.cell(row=7, column=56).value
			#入札者認証制限を設定する
			limitAuth = yafuokuSheet.cell(row=7, column=57).value
			#出品者情報の手動開示
			manualDisclosure = yafuokuSheet.cell(row=7, column=58).value
			#最低落札価格
			lowestWinningBid = yafuokuSheet.cell(row=7, column=59).value
			#注目のオークション
			attentionAuction = yafuokuSheet.cell(row=7, column=60).value
			#あなたへのおすすめコレクション
			recomendCollection = yafuokuSheet.cell(row=7, column=61).value
			#アフィリエイト
			affiliate = yafuokuSheet.cell(row=7, column=62).value
			#太字テキスト
			boldText = yafuokuSheet.cell(row=7, column=63).value
			#背景色
			backgroundColor = yafuokuSheet.cell(row=7, column=64).value
			#目立ちアイコン
			attentionIcon = yafuokuSheet.cell(row=7, column=65).value
			#贈答品アイコン
			giftIcon = yafuokuSheet.cell(row=7, column=66).value
			#みんなのチャリティー
			ourCharity = yafuokuSheet.cell(row=7, column=67).value



			#商品名
			if goodsName == None:
				print("商品名をスキップ")
			else:
				driver.find_element_by_xpath("//input[@id='fleaTitleForm']").send_keys(goodsName)

			#カテゴリ
			driver.find_element_by_id('acMdCateChange').send_keys(Keys.ENTER)
			sleep(2)
			driver.find_element(
				By.XPATH, '//div[@class="Tab__itemIn"]').find_element(By.XPATH, '..').send_keys(Keys.ENTER)
			sleep(2)

			#指定されたフレームが利用出来るまで待機する
			WebDriverWait(driver, 500).until(
				EC.presence_of_element_located((By.CLASS_NAME, "decSlctList"))
			)
			sleep(5)

			category = 'コンピュータ'
			xpath = "//a[text()='%s']" % category
			driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
			sleep(2)

			category = 'パソコン'
			xpath = "//a[text()='%s']" % category
			driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
			sleep(2)

			category = 'Windows'
			xpath = "//a[text()='%s']" % category
			driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
			sleep(2)

			cc = 1
			isMatchCategory = False
			#カテゴリシートの行番号でループ				
			while cc <= yafuokuCategorySheet.max_row:
				categoryString = yafuokuCategorySheet.cell(row=cc, column=1).value
				categoryString2 = yafuokuCategorySheet.cell(row=cc, column=2).value

				#A列のカテゴリとB列のカテゴリが一致する場合はcategory2とcategory3が確定
				if categoryString == goodsSheet.cell(row=rc, column=21).value:
					if categoryString2 == goodsSheet.cell(row=rc, column=6).value:
						isMatchCategory = True
						break
				cc = cc + 1

			xpath = "//a[text()='%s']" % goodsSheet.cell(row=rc, column=21).value
			driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
			sleep(2)

			if isMatchCategory == True:
				category2 = yafuokuCategorySheet.cell(row=cc, column=3).value
				category3 = yafuokuCategorySheet.cell(row=cc, column=4).value

				xpath = "//a[text()='%s']" % category2
				driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
				sleep(2)

				xpath = "//a[text()='%s']" % category3
				driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
				sleep(2)

			else:
				#その他を選択
				list4Element = driver.find_element_by_id("ptsSlctList4")
				#.//は子要素検索
				xpath = ".//a[text()='%s']" % "その他"
				list4Element.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
				sleep(2)

			#カテゴリが確定した場合のみcategory3はクリック可能
			if isMatchCategory == True:
				xpath = "//a[text()='%s']" % category3
				driver.find_element(By.XPATH, xpath).send_keys(Keys.ENTER)
				sleep(2)

			driver.find_element(By.XPATH, '//span[text()="このカテゴリに出品"]').find_element(By.XPATH,'..').send_keys(Keys.ENTER)
			sleep(2)

			#商品の状態
			if status != None:
				Select(driver.find_element_by_xpath("//select[@name='istatus']")).select_by_visible_text(status)
				sleep(2)

			#商品説明
			if description_html != None:
				driver.find_element(By.XPATH, '//*[@id="aucHTMLtag"]').send_keys(Keys.ENTER)
				sleep(2)
				pyperclip.copy(description_html)
				driver.find_element(By.XPATH, '//*[@id="textMode"]/div[2]/textarea[1]').send_keys(Keys.CONTROL+ "v")
				sleep(2)

			#個数
			#Select(driver.find_element_by_xpath("//select[@name='Quantity']")).select_by_visible_text("1個")
			#sleep(2)

			#発送元の地域
			Select(driver.find_element_by_xpath(
				"//select[@name='loc_cd']")).select_by_visible_text(destinationRegion)
			sleep(2)

			#配送の負担人
			Select(driver.find_element_by_xpath(
				"//select[@id='auc_shipping_who']")).select_by_visible_text(postageBurdenPerson)
			sleep(2)

			#落札者を選択した場合(送料を決めずに出品する場合のクリックが必要)
			if postageBurdenPerson == "落札者":
				if shippingCostWay != "-":
					element = driver.find_element_by_xpath("//dt[@id='auc_shipping_option']")
					driver.execute_script("arguments[0].click();", element)

					if shippingCostWay == "落札された後に送料を連絡する":
						element = driver.find_element(By.XPATH, '//input[@id="later_check"]').find_element(By.XPATH, '..')
						driver.execute_script("arguments[0].click();", element)
					elif shippingCostWay == "着払いにする":
						element = driver.find_element(By.XPATH, '//input[@id="arrival_check"]').find_element(By.XPATH, '..')
						driver.execute_script("arguments[0].click();", element)

			#配送方法
			#一旦すべてのチェックボックスをクリア
			els = driver.find_elements(By.XPATH, '//label[@class="CheckExpand__label cf  is-check"]')
			for el in els:
				driver.execute_script("arguments[0].click();", el)

			sleep(2)

			# ヤフネコ!パック
			if postageBurdenPerson != "落札者" or  shippingCostWay != "着払いにする":
					#ヤフネコ！ネコポス
					if yafunekoNekopos != "-":
						driver.find_element_by_xpath( "//input[@id='ship_delivery_n']" ).find_element(By.XPATH,'..').click()
						sleep(2)

					#ヤフネコ！宅配便コンパクト
					if yafunekoTakuConpact != "-":
						driver.find_element_by_xpath( "//input[@id='ship_delivery_s']" ).find_element(By.XPATH,'..').click()
						sleep(2)

					#ヤフネコ！宅配便
					if yafunekoTaku != "-":
						driver.find_element_by_xpath( "//input[@id='ship_delivery_l']" ).find_element(By.XPATH,'..').click()
						sleep(2)
						if postageBurdenPerson == "落札者" and shippingCostWay == "-":
							Select(driver.find_element(By.XPATH, '//select[@id="ship_delivery_l_size_select"]')).select_by_visible_text(yafunekoTakuSize)
							sleep(2)
							Select(driver.find_element(By.XPATH, '//select[@id="ship_delivery_l_weight_select"]')).select_by_visible_text(yafunekoTakuWeight)
							sleep(2)

					#ゆうパック・ゆうパケット（おてがる版）
					#ゆうパケット
					if yupaketOtegaru != "-":
						driver.find_element_by_xpath( "//input[@id='ship_delivery_yupacket']" ).find_element(By.XPATH,'..').click()
						sleep(2)

					#ゆうパック
					if yupackOtegaru != "-":
						driver.find_element_by_xpath( "//input[@id='ship_delivery_yupack']" ).find_element(By.XPATH,'..').click()
						sleep(2)
						if postageBurdenPerson == "落札者" and shippingCostWay == "-":
							Select(driver.find_element(By.XPATH, '//select[@id="ship_delivery_yupack_size_select"]')).select_by_visible_text(yupackSize)
							sleep(2)
							Select(driver.find_element(By.XPATH, '//select[@id="ship_delivery_yupack_weight_select"]')).select_by_visible_text(yupackWeight)
							sleep(2)

			#そのほかの郵送方法１
			if sendMethodOther1 != "-":
				#追加ボタンを押す
				element = driver.find_element_by_xpath("//span[@id='auc_add_shipform']")
				driver.execute_script("arguments[0].click();", element)
				sleep(2)
				#チェックボックス
				if driver.find_element_by_xpath("//input[@id='shipping_other_check1']").is_selected() == False:
					element = driver.find_element_by_xpath("//input[@id='shipping_other_check1']").find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)
				Select(driver.find_element_by_xpath("//select[@id='auc_shipname_standard1']")).select_by_visible_text(sendMethodOther1)
				sleep(2)

				feeUniform = 0
				feeHokkaido = 0
				feeOkinawa =0
				feeIsland = 0

				if categoryString == "デスクトップ":
					feeUniform = 2000
					feeHokkaido = 2500
					feeOkinawa = 2500
					feeIsland =  2500         
				elif categoryString == "ノートブック、ノートパソコン":
					feeUniform = 1500
					feeHokkaido = 2000
					feeOkinawa = 2500
					feeIsland = 2500
				else:
					feeUniform = 2500
					feeHokkaido = 3000
					feeOkinawa = 3500
					feeIsland = 3500

				#送料入力有の場合
				if shippingCostWay == "-":
					if isSetFee == 1:
						driver.find_element_by_xpath("//input[@id='auc_shipname_uniform_fee_data1']").clear()
						driver.find_element_by_xpath("//input[@id='auc_hokkaidoshipping1']").clear()
						driver.find_element_by_xpath("//input[@id='auc_okinawashipping1']").clear()
						driver.find_element_by_xpath("//input[@id='auc_isolatedislandshipping1']").clear()
						sleep(2)
						driver.find_element_by_xpath("//input[@id='auc_shipname_uniform_fee_data1']").send_keys(feeUniform)
						driver.find_element_by_xpath("//input[@id='auc_hokkaidoshipping1']").send_keys(feeHokkaido)
						driver.find_element_by_xpath("//input[@id='auc_okinawashipping1']").send_keys(feeOkinawa)
						driver.find_element_by_xpath("//input[@id='auc_isolatedislandshipping1']").send_keys(feeIsland)

			#支払いから発送までの日数
			Select(driver.find_element_by_xpath("//select[@name='shipschedule']")).select_by_visible_text(sendToDate)
			sleep(2)

			#販売形式
			if sellingAuction != "-":					
				element = driver.find_elements(By.XPATH, '//input[@name="salesmode"]')[0].find_element(By.XPATH, '..')
				driver.execute_script("arguments[0].click();", element)
				sleep(2)
				if auctionStartPrice != "-":
					driver.find_element(By.XPATH, '//input[@name="StartPrice"]').send_keys(int(auctionStartPrice))
					sleep(2)
				if auctionDecidePrice != "-":
					element = driver.find_element(By.XPATH, "//*[@id='price_auction']/div[2]/dl/dt")
					driver.execute_script("arguments[0].click();", element)
					driver.find_element(By.XPATH, '//input[@name="BidOrBuyPrice"]').send_keys(int(auctionDecidePrice))
					sleep(2)
			else:
				element = driver.find_elements(By.XPATH, '//input[@name="salesmode"]')[1].find_element(By.XPATH, '..')
				driver.execute_script("arguments[0].click();", element)
				driver.find_element(By.XPATH, '//input[@name="BidOrBuyPrice"]').send_keys(int(auctionDecidePrice))
				driver.find_element(By.XPATH, '//input[@id="auc_BidOrBuyPrice_buynow"]').send_keys(int(fixedPrice))
				sleep(2)
				if reductionNego != "-":
					element = driver.find_element(By.XPATH, '//input[@id="salesmode_offer"]').find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)

			#終了する日
			Select(driver.find_element_by_xpath("//select[@name='ClosingYMD']")).select_by_visible_text(endDay)
			sleep(2)

			#終了する時間
			Select(driver.find_element_by_xpath("//select[@name='ClosingTime']")).select_by_visible_text(endHour)
			sleep(2)

			# 自動再出品のトグルボタンを押す(押されていない場合)
			try:
				Select(driver.find_element_by_xpath('//select[@name="numResubmit"]')).select_by_visible_text(autoReSellingCount)
			except:
				xpath = "//*[text()='%s']" % "自動再出品を設定する"
				element = driver.find_element(By.XPATH, xpath)
				driver.execute_script("arguments[0].click();", element)
				pass

			# 自動再出品設定(回数)
			if autoPostponed == "○":
				Select(driver.find_element_by_xpath('//select[@name="numResubmit"]')).select_by_visible_text(autoReSellingCount)


			#　ここからオプションの設定----------------------------------------------------------------------------------
			# オプションのトグルボタンを押す(押されていない場合)
			try:
				element = driver.find_element(By.XPATH, '//input[@name="AutoExtension"]').find_element(By.XPATH, '..')
				driver.execute_script("arguments[0].click();", element)
			except:
				element = driver.find_element_by_xpath("//*[@id='option_Area']/div[1]/div")
				driver.execute_script("arguments[0].click();", element)
				pass

			#終了時間を自動延長する
			if autoPostponed == "○":					
				if driver.find_element(By.XPATH, '//input[@name="AutoExtension"]').is_selected() == False:
					element = driver.find_element(By.XPATH, '//input[@name="AutoExtension"]').find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)
			else:
				if driver.find_element(By.XPATH, '//input[@name="AutoExtension"]').is_selected() == True:
					element = driver.find_element(By.XPATH, '//input[@name="AutoExtension"]').find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)

			#返品を受け取る
			if getReturn == "○":
				if driver.find_elements(By.XPATH, '//input[@name="retpolicy"]')[1].is_selected() == False:
					element = driver.find_elements(By.XPATH, '//input[@name="retpolicy"]')[1].find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)
			else:
				if driver.find_elements(By.XPATH, '//input[@name="retpolicy"]')[1].is_selected() == True:
					element = driver.find_elements(By.XPATH, '//input[@name="retpolicy"]')[1].find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)

			#出品者情報を手動で開示する
			if autoPostponed == "○":					
				if driver.find_element(By.XPATH, '//input[@name="salesContract"]').is_selected() == False:
					element = driver.find_element(By.XPATH, '//input[@name="salesContract"]').find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)
			else:
				if driver.find_element(By.XPATH, '//input[@name="salesContract"]').is_selected() == True:
					element = driver.find_element(By.XPATH, '//input[@name="salesContract"]').find_element(By.XPATH, '..')
					driver.execute_script("arguments[0].click();", element)
					sleep(2)


			#element = driver.find_element(By.XPATH, '//*[@id="modFormReqrd"]/ul/li[2]/input')
			element = driver.find_element(By.XPATH, "//input[@value='確認する']")
			driver.execute_script("arguments[0].click();", element)
			sleep(2)

			#出品するかしないか
			if isExhibitString == "T":					
				element = driver.find_element(By.XPATH, '//*[@id="auc_preview_submit_up"]')
				driver.execute_script("arguments[0].click();", element)
				sleep(2)

		except Exception as yafuoku_err:
			logger.exception('Raise Exception yafuoku: %s', yafuoku_err)

		# 処理開始 ----------------------ジモティー
		try:
			isSellingJmty = True
			logger.info("jmty" + str(count))
			driver.execute_script("window.open()") #新しいタブを開く
			# ウィンドウハンドルを取得する
			handle_array = driver.window_handles

			# 一番最後のdriverに切り替える
			driver.switch_to.window(handle_array[len(handle_array)-1])

			# ジモティーにアクセス
			driver.get('https://jmty.jp/articles/new?prefecture_id=0')
			#指定されたフレームが利用出来るまで待機する
			WebDriverWait(driver, 500).until(
				EC.presence_of_element_located((By.ID, "article_form"))
			)

			# 読み込み遅いかもしれないから3秒待つ。
			sleep(3)

			#画像を設定--------------
			import glob
			folderPath = "input_data/image/" + str(goodsSheet.cell(row=rc, column=22).value)
			files = glob.glob(folderPath + "/*" )
			imgCount = 1
			for file_name in files:								
				if os.path.exists(file_name):						
					xpath = '//a[@id="upload_link"]'
					element = driver.find_element(By.XPATH, xpath).click()
					sleep(2)

					import pywinauto
					# 開くダイアログを探して接続する
					findWindow = lambda: pywinauto.findwindows.find_windows(title=u'開く')[0]
					dialog = pywinauto.timings.wait_until_passes(5, 1, findWindow)

					pwa_app = pywinauto.Application()
					pwa_app.connect(handle=dialog)

					# pywinauto に探し出したダイアログを接続
					pwa_app = pywinauto.Application()
					pwa_app.connect(handle=dialog)
					window = pwa_app[u"開く"]

					addres = window.children()[39]
					addres.click()

					dialog_dir = window.children()[43]
					dialog_dir.type_keys( dirPath +"\\"+ folderPath+'{ENTER}',with_spaces=True)

					filePath = file_name.rsplit('\\', 1)[1]
					# テキストボックス(ファイル名)にPATHを入力
					tb = window[u"ファイル名(&N):"]
					if tb.is_enabled():
						tb.click()
						edit = window.Edit4
						edit.set_focus()
						# ファイルを選択し、Alt + Oを押下
						edit.type_keys(filePath + '%O',with_spaces=True)

					sleep(3)
					imgCount = imgCount +1   
					if imgCount > 5:
						break

			# エクセルファイルから値を取得-----------
			#カテゴリ
			category1 = "売ります・あげます"
			category2 = "パソコン"
			# カテゴリ３はデスクトップかノートパソコンかの分岐しかない
			category3 = "デスクトップ"
			if goodsSheet.cell(row=rc, column=21).value == "ノートブック、ノートパソコン":
				category3 = "ノートパソコン"
			category4 = "-"

			#支払い方法は「現金」固定
			#paymentMethod 

			#問い合わせ（質問など）を受け付ける 
			contact = jmtySheet.cell(row=7, column=16).value

			#出品エリア
			prefecture = jmtySheet.cell(row=7, column=17).value 
			city1 = jmtySheet.cell(row=7, column=18).value 
			city2 = jmtySheet.cell(row=7, column=19).value
			station1 = jmtySheet.cell(row=7, column=20).value 
			station2 = jmtySheet.cell(row=7, column=21).value 

			#受け渡し方法は「配送」固定
			#deliveryMethod

			#配送料
			deliveryFee = jmtySheet.cell(row=7, column=23).value

			#商品価格
			price = auctionStartPrice

			# 入力処理--------------------

			#カテゴリー
			if not category1:
				pass
			else:
				if category1 == "-":
					pass
				else:
					Select(driver.find_element_by_xpath( "//select[@id='category_group_id']")).select_by_visible_text(category1)
					sleep(2)

			if not category2:
				pass
			else:
				if category2 == "-":
					pass
				else:
					Select(driver.find_element_by_xpath( "//select[@id='article_category_id']")).select_by_visible_text(category2)
					sleep(2)

			if not category3:
				pass
			else:
				if category3 == "-":
					pass
				else:
					Select(driver.find_element_by_xpath( "//select[@id='article_large_genre_id']")).select_by_visible_text(category3)
					sleep(2)

			if not category4:
				pass
			else:
				if category4 == "-":
					pass
				else:
					Select(driver.find_element_by_xpath( "//select[@id='article_medium_genre_id']")).select_by_visible_text(category4)
					sleep(2)


			#タイトル
			driver.find_element_by_xpath( "//input[@id='article_title']" ).send_keys(goodsName)
			sleep(2)

			#内容
			pyperclip.copy(description_text)
			driver.find_element_by_xpath( "//textarea[@id='article_text']" ).send_keys(Keys.CONTROL+ "v")
			sleep(2)

			#出品エリア
			Select(driver.find_element_by_xpath("//select[@id='article_prefecture_id']")).select_by_visible_text(prefecture)
			sleep(2)

			Select(driver.find_element_by_xpath("//select[@id='article_city_id']")).select_by_visible_text(city1)
			sleep(2)

			Select(driver.find_element_by_xpath("//select[@id='article_town_id']")).select_by_visible_text(city2)
			sleep(2)

			Select(driver.find_element_by_xpath("//select[@id='article_line_id1']")).select_by_visible_text(station1)
			sleep(2)

			Select(driver.find_element_by_xpath("//select[@id='article_station_master_id1']")).select_by_visible_text(station2)
			sleep(2)

			#問い合わせ チェック
			if driver.find_element(By.XPATH, "//input[@id='article_online_purchasable_inquirable']").is_selected() == False:
				driver.find_element_by_xpath( "//input[@id='article_online_purchasable_inquirable']" ).click()
				sleep(2)

			#受け渡し方法
			#配送にチェック
			if driver.find_element(By.XPATH, "//input[@id='article_delivery_option_attributes_by_seller']").is_selected() == False:
				driver.find_element_by_xpath( "//input[@id='article_delivery_option_attributes_by_seller']" ).click()
				sleep(2)

			#配送料
			driver.find_element_by_xpath( "//input[@id='article_delivery_option_attributes_seller_carriage']" ).clear()
			sleep(1)
			driver.find_element_by_xpath( "//input[@id='article_delivery_option_attributes_seller_carriage']" ).send_keys(deliveryFee)
			sleep(2)

			#商品価格
			driver.find_element_by_xpath( "//input[@id='article_price']" ).clear()
			sleep(1)
			driver.find_element_by_xpath( "//input[@id='article_price']" ).send_keys(price)
			sleep(2)

			#出品するかしないか
			if isExhibitString == "T":
				xpath = '//input[@id="article_submit_button"]'
				element = driver.find_element(By.XPATH, xpath)
				element.send_keys(Keys.ENTER)

		except Exception as jmty_err:
			logger.exception('Raise Exception jmty: %s', jmty_err)



		#出品後フォルダ/today(yyyyMMdd)へ移動する
		path_dir = "input_data/image/" + str(goodsSheet.cell(row=rc, column=22).value)
		if os.path.exists(path_dir):
    		#日付フォルダが存在しない場合は作成する
			today = datetime.date.today()
			move_dir = LISTED_DIR + "\\" + today.strftime('%Y%m%d')
			if not os.path.exists(move_dir):
    			# ディレクトリが存在しない場合、ディレクトリを作成する
				os.makedirs(move_dir)
			shutil.move(path_dir,move_dir)

		#自動出品かつ待ち時間指定がある場合は、ここで
		if isExhibitString =="T":
			logger.exception("waitting exhibitWaitTime")
			wt = random.randint(1,5) + exhibitWaitTime
			sleep(wt)

		rc = rc +1

except Exception as err:
	logger.exception('Raise Exception: %s', err)
	returnCode= -1

# プログラム終了
logger.info("Auto Selling Pc All End")
sys.exit(returnCode)